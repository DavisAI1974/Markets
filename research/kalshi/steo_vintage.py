"""
steo_vintage.py - FEED T (family D/balance): STEO monthly VINTAGES - the frozen as-of NG balance (S99).

WHY THIS EXISTS
---------------
The free weekly balance LEVELS died 2025-10-02 (feed N, measured), so the walked winter has no free
weekly S/D levels. The S98 EIA sweep (EIA_BALANCE_OPTIONS_S98.md) found the strongest free
substitute: the Short-Term Energy Outlook publishes the COMPLETE U.S. natural gas balance every
month - dry production, marketed production by basin, consumption by sector, LNG + pipeline trade,
net storage withdrawals, end-of-month working inventory by region - and EIA ARCHIVES EVERY MONTHLY
VINTAGE as a frozen workbook. Table 5a of each archive workbook is the balance as-of that release.
This feed puts that as-of read - at its honest 1-34 day staleness - in front of the agent. The
vintage-to-vintage evolution (e.g. the Jan-13 -> Feb-10 pair bracketing the January freeze re-mark:
January consumption 115.95 -> 121.90 Bcf/d, end-Jan working gas 2,609 -> 2,472 Bcf) is itself
signal, the same first-appearance-vs-revision structure the MOS feed carries. The feed does not
gate, score or recommend; the agent decides what it means.

THE BLIND WALL (measured, S98 sweep section 1b)
-----------------------------------------------
Official release dates were measured from Wayback captures of the live STEO landing page (the dates
visitors actually saw). Releases are ~noon ET, so the join convention is
    knowable_from = official_release_date + 1 calendar day
(matches feeds G/N). THE NAMED TRAP: the archive files' HTTP Last-Modified stamps equal the
forecast-COMPLETION date, 3-6 days BEFORE public release, on four of seven vintages -
Last-Modified is file-write time, NOT availability evidence. This module joins ONLY on the
measured release dates below and never reads Last-Modified.

    vintage   completed   RELEASED      note
    sep25     2025-09-04  2025-09-09
    oct25     2025-10-02  2025-10-07    published DURING the shutdown
    nov25     2025-11-06  2025-11-12    slipped ~1 week; landed the day the shutdown ended
    dec25     2025-12-04  2025-12-09
    jan26     2026-01-08  2026-01-13
    feb26     2026-02-05  2026-02-10    carries the post-freeze re-mark
    mar26     2026-03-09  2026-03-10

SCOPE: these seven vintages cover every walked-winter decision day (first knowable 2025-09-10) plus
G12/G13. Later vintages (apr26+) need their OWN measured release dates before joining - a
live-forward extension, deliberately not assumed here.

HONEST LABELING (sweep 1d): values for months at/after the NGM measurement anchor are STIFS MODEL
ESTIMATES (anchored on indicator data), published by EIA as-printed. Each issue is a frozen file,
so the as-printed record has ZERO vintage risk by construction. The history/estimate per-cell break
(gray shading) is not parsed; the label rides on the whole block.

PARSE TRAP (hit and solved in the sweep): the monthly column origin CHANGES between vintages
(2021-01 for sep25-dec25, 2022-01 for jan26+). The parser detects the header row / month mapping
per workbook and never assumes an origin.

STORE: data/steo_vintage/ (raw workbooks + steo_vintage.json.gz). ALL series rows of Table 5a are
kept (raw-ingestion discipline); the asof exposes a curated subset plus revision deltas, and the
full store stays accessible via load_store(). Missing is None, never zero. No commits by this
module; S3 push is the orchestrator's step (prefix steo_vintage/).

USAGE
-----
  python research/kalshi/steo_vintage.py --build       # download + parse + write store
  python research/kalshi/steo_vintage.py --selftest
  python research/kalshi/steo_vintage.py --show 2026-02-01
  python research/kalshi/steo_vintage.py --list-ids    # all series IDs in the store
"""

from __future__ import annotations

import argparse
import datetime
import gzip
import json
import os
import sys

_ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
STORE_DIR = os.path.join(_ROOT, "data", "steo_vintage")
RAW_DIR = os.path.join(STORE_DIR, "raw")
STORE_PATH = os.path.join(STORE_DIR, "steo_vintage.json.gz")

ARCHIVE_URL = "https://www.eia.gov/outlooks/steo/archives/{vid}_base.xlsx"
_UA = ("Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
       "(KHTML, like Gecko) Chrome/126.0 Safari/537.36")

# MEASURED release dates - Wayback captures of the STEO landing page (EIA_BALANCE_OPTIONS_S98.md 1b).
# NEVER extend this table from Last-Modified stamps or assumptions; measure first.
VINTAGES = [
    # (vintage_id, forecast_completed, official_release_date)
    ("sep25", "2025-09-04", "2025-09-09"),
    ("oct25", "2025-10-02", "2025-10-07"),
    ("nov25", "2025-11-06", "2025-11-12"),
    ("dec25", "2025-12-04", "2025-12-09"),
    ("jan26", "2026-01-08", "2026-01-13"),
    ("feb26", "2026-02-05", "2026-02-10"),
    ("mar26", "2026-03-09", "2026-03-10"),
]

MONTH_ABBR = {"Jan": 1, "Feb": 2, "Mar": 3, "Apr": 4, "May": 5, "Jun": 6,
              "Jul": 7, "Aug": 8, "Sep": 9, "Oct": 10, "Nov": 11, "Dec": 12}

# Curated exposure set: field -> candidate series IDs (first present wins; the ID used is recorded
# per read). Everything else in Table 5a stays in the store, accessible via load_store().
CURATED = [
    ("dry_production_bcfd",        ["NGPRPUS"]),
    ("marketed_production_bcfd",   ["NGMPPUS"]),
    ("total_consumption_bcfd",     ["NGTCPUS"]),
    ("residential_bcfd",           ["NGRCPUS"]),
    ("commercial_bcfd",            ["NGCCPUS"]),
    ("industrial_bcfd",            ["NGINPUS", "NGINCON"]),
    ("power_burn_bcfd",            ["NGEPCON", "NGEPPUS"]),
    ("lng_gross_exports_bcfd",     ["NGEXPUS_LNG"]),
    ("lng_gross_imports_bcfd",     ["NGIMPUS_LNG"]),
    ("pipeline_gross_exports_bcfd", ["NGEXPUS_PIPE"]),
    ("pipeline_gross_imports_bcfd", ["NGIMPUS_PIPE"]),
    ("net_storage_withdrawal_bcfd", ["NGNWPUS"]),
    ("working_gas_inventory_bcf",  ["NGWGPUS"]),
]


def _iso_plus_days(iso: str, n: int) -> str:
    return (datetime.date.fromisoformat(iso) + datetime.timedelta(days=n)).isoformat()


def _download(vid: str, force: bool = False) -> str:
    """GET one archive workbook (browser UA - the bare listing 403s, the files serve fine)."""
    import requests
    os.makedirs(RAW_DIR, exist_ok=True)
    path = os.path.join(RAW_DIR, f"{vid}_base.xlsx")
    if os.path.exists(path) and not force:
        return path
    url = ARCHIVE_URL.format(vid=vid)
    r = requests.get(url, headers={"User-Agent": _UA}, timeout=120)
    r.raise_for_status()
    if len(r.content) < 200_000:
        raise RuntimeError(f"{vid}: suspiciously small response ({len(r.content)} bytes) - not a workbook?")
    with open(path, "wb") as f:
        f.write(r.content)
    return path


def _find_5a_sheet(wb):
    for name in wb.sheetnames:
        if name.lower() == "5atab":
            return wb[name]
    for name in wb.sheetnames:
        if "5a" in name.lower():
            return wb[name]
    raise RuntimeError(f"no Table 5a sheet found; sheets={wb.sheetnames}")


def _month_key(y: int, m: int) -> str:
    return f"{y:04d}{m:02d}"


def _header_map(ws) -> dict[int, str]:
    """Detect the monthly header columns. Strategy A: a row of datetime cells. Strategy B: a
    month-abbreviation row with a year row above it (merged years forward-filled). The column
    origin CHANGES between vintages - always detected, never assumed."""
    rows = list(ws.iter_rows(min_row=1, max_row=10, values_only=True))
    # Strategy A: datetimes in one row
    best_row, best_cols = None, {}
    for ri, row in enumerate(rows):
        cols = {ci: _month_key(v.year, v.month) for ci, v in enumerate(row)
                if isinstance(v, (datetime.datetime, datetime.date))}
        if len(cols) > len(best_cols):
            best_row, best_cols = ri, cols
    if len(best_cols) >= 12:
        return best_cols
    # Strategy B: month-name row + year row above
    for ri, row in enumerate(rows):
        mcols = {ci: str(v).strip() for ci, v in enumerate(row)
                 if isinstance(v, str) and str(v).strip()[:3] in MONTH_ABBR and len(str(v).strip()) <= 4}
        if len(mcols) < 12:
            continue
        for yi in range(ri - 1, -1, -1):
            yrow = rows[yi]
            years = {ci: int(v) for ci, v in enumerate(yrow)
                     if isinstance(v, (int, float)) and not isinstance(v, bool) and 2015 <= int(v) <= 2035}
            if not years:
                continue
            cols, cur_year = {}, None
            for ci in sorted(mcols):
                if ci in years:
                    cur_year = years[ci]
                else:  # merged year cell: forward-fill from the last seen year boundary
                    prior = [c for c in years if c <= ci]
                    if prior:
                        cur_year = years[max(prior)]
                m = MONTH_ABBR[mcols[ci][:3]]
                if cur_year is not None:
                    # a January column after the fill point advances the year when months wrap
                    cols[ci] = _month_key(cur_year, m)
            if len(cols) >= 12:
                # fix year wrap for merged-cell fills: months must be monotone by column
                fixed, last = {}, None
                year_adjust = 0
                for ci in sorted(cols):
                    ym = cols[ci]
                    y, m = int(ym[:4]) + year_adjust, int(ym[4:])
                    if last is not None and (y, m) <= last:
                        y += 1
                        year_adjust += 1
                    fixed[ci] = _month_key(y, m)
                    last = (y, m)
                return fixed
    raise RuntimeError("could not detect the monthly header row (neither datetime nor month-name form)")


def _parse_workbook(path: str) -> dict:
    """Parse Dates sheet (raw label:value pairs, informational) + Table 5a (ALL series rows kept)."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    dates_raw = {}
    for name in wb.sheetnames:
        if name.lower() == "dates":
            for row in wb[name].iter_rows(min_row=1, max_row=40, values_only=True):
                vals = [v for v in row if v is not None]
                if len(vals) >= 2 and isinstance(vals[0], str):
                    dates_raw[vals[0].strip()] = str(vals[1])
            break
    ws = _find_5a_sheet(wb)
    hdr = _header_map(ws)
    series: dict[str, dict] = {}
    for row in ws.iter_rows(values_only=True):
        rid = row[0] if row else None
        if not isinstance(rid, str) or not rid.strip():
            continue
        rid = rid.strip()
        vals = {hdr[ci]: round(float(v), 4) for ci, v in enumerate(row)
                if ci in hdr and isinstance(v, (int, float)) and not isinstance(v, bool)}
        if not vals:
            continue
        desc = next((str(v).strip() for v in row[1:4] if isinstance(v, str) and str(v).strip()), None)
        if rid in series:  # duplicate ID: keep first, name the event
            series[rid]["dup_count"] = series[rid].get("dup_count", 1) + 1
            continue
        series[rid] = {"desc": desc, "values": vals}
    wb.close()
    months = sorted({m for s in series.values() for m in s["values"]})
    if not series:
        raise RuntimeError(f"{path}: no series rows parsed")
    return {"dates_sheet_raw": dates_raw, "series": series,
            "column_origin": months[0], "column_end": months[-1], "n_series": len(series)}


def build(force: bool = False) -> dict:
    os.makedirs(STORE_DIR, exist_ok=True)
    vintages = {}
    for vid, completed, released in VINTAGES:
        path = _download(vid, force=force)
        parsed = _parse_workbook(path)
        vintages[vid] = {
            "forecast_completed": completed,
            "release_date": released,
            "knowable_from": _iso_plus_days(released, 1),
            "raw_file": os.path.relpath(path, _ROOT),
            **parsed,
        }
        print(f"[steo_vintage] {vid}: {parsed['n_series']} series, cols {parsed['column_origin']}"
              f"..{parsed['column_end']}, released {released} (knowable {vintages[vid]['knowable_from']})")
    store = {
        "meta": {
            "built_utc": datetime.datetime.now(datetime.timezone.utc).isoformat(timespec="seconds"),
            "source": ARCHIVE_URL,
            "release_dates_provenance": "Wayback captures of the STEO landing page, measured "
                                        "2026-07-20 (EIA_BALANCE_OPTIONS_S98.md section 1b); "
                                        "Last-Modified is NOT availability evidence and is never used",
            "label": "values at/after the NGM anchor are STIFS model estimates, as-printed, "
                     "frozen per issue (zero vintage risk by construction)",
            "scope": "seven walked-winter vintages sep25..mar26; apr26+ needs its own measured "
                     "release dates before joining",
        },
        "vintages": vintages,
    }
    with gzip.open(STORE_PATH, "wt", encoding="utf-8") as f:
        json.dump(store, f)
    print(f"[steo_vintage] store written: {os.path.relpath(STORE_PATH, _ROOT)} "
          f"({len(vintages)} vintages)")
    return store


_CACHE: dict | None = None


def load_store() -> dict | None:
    global _CACHE
    if _CACHE is None and os.path.exists(STORE_PATH):
        with gzip.open(STORE_PATH, "rt", encoding="utf-8") as f:
            _CACHE = json.load(f)
    return _CACHE


def _first_present(series: dict, candidates: list[str]):
    for cid in candidates:
        if cid in series:
            return cid
    return None


def _shift_month(ym: str, k: int) -> str:
    y, m = int(ym[:4]), int(ym[4:])
    t = y * 12 + (m - 1) + k
    return _month_key(t // 12, t % 12 + 1)


def steo_vintage_asof(iso: str) -> dict | None:
    """The latest STEO vintage knowable at date iso (knowable_from = release+1), with the curated
    balance snapshot for prev/current/next month and the revision deltas vs the PRIOR vintage.
    None before the first knowable vintage. Missing months/series are None, never zero."""
    store = load_store()
    if not store:
        return None
    known = [(v["release_date"], vid, v) for vid, v in store["vintages"].items()
             if v["knowable_from"] <= iso]
    if not known:
        return None
    known.sort()
    _, vid, v = known[-1]
    order = [x[0] for x in VINTAGES]
    prev_vid = order[order.index(vid) - 1] if order.index(vid) > 0 else None
    prev = store["vintages"].get(prev_vid) if prev_vid else None
    cur_m = iso[:4] + iso[5:7]
    months = {"prev": _shift_month(cur_m, -1), "cur": cur_m, "next": _shift_month(cur_m, 1)}
    fields, revisions = {}, {}
    for fname, cands in CURATED:
        cid = _first_present(v["series"], cands)
        if cid is None:
            fields[fname] = None
            continue
        vals = v["series"][cid]["values"]
        fields[fname] = {"series_id": cid,
                         **{slot: vals.get(m) for slot, m in months.items()}}
        if prev and cid in prev["series"]:
            pv = prev["series"][cid]["values"]
            deltas = {slot: (round(vals[m] - pv[m], 4) if (m in vals and m in pv) else None)
                      for slot, m in months.items()}
            if any(d is not None for d in deltas.values()):
                revisions[fname] = deltas
    rel = datetime.date.fromisoformat(v["release_date"])
    return {
        "vintage_id": vid,
        "release_date": v["release_date"],
        "knowable_from": v["knowable_from"],
        "age_days": (datetime.date.fromisoformat(iso) - rel).days,
        "month_prev": months["prev"], "month_cur": months["cur"], "month_next": months["next"],
        "fields": fields,
        "revisions_vs_prev_vintage": revisions or None,
        "prev_vintage_id": prev_vid,
        "label": "STIFS estimates at/after the NGM anchor; as-printed, frozen per issue",
    }


def _t(cond, msg, fails):
    print(("  PASS " if cond else "  FAIL ") + msg)
    return fails + (0 if cond else 1)


def _selftest() -> int:
    print("=== steo_vintage --selftest ===")
    fails = 0
    store = load_store()
    fails = _t(store is not None and len(store["vintages"]) == 7, "store holds 7 vintages", fails)
    if store is None:
        return 1
    # value pins from the S98 sweep's extracted vintage tables (EIA_BALANCE_OPTIONS_S98.md 1c)
    sv = store["vintages"]
    def val(vid, sid, ym):
        s = sv[vid]["series"].get(sid)
        return s["values"].get(ym) if s else None
    pins = [
        ("sep25", "NGPRPUS", "202509", 106.68), ("oct25", "NGPRPUS", "202511", 108.17),
        ("dec25", "NGPRPUS", "202512", 110.31), ("jan26", "NGPRPUS", "202601", 109.22),
        ("feb26", "NGPRPUS", "202601", 106.68), ("mar26", "NGPRPUS", "202602", 110.09),
        ("nov25", "NGTCPUS", "202511", 93.09), ("jan26", "NGTCPUS", "202601", 115.95),
        ("feb26", "NGTCPUS", "202601", 121.90), ("jan26", "NGEXPUS_LNG", "202601", 17.08),
    ]
    for vid, sid, ym, exp in pins:
        got = val(vid, sid, ym)
        fails = _t(got is not None and abs(got - exp) < 0.01,
                   f"{vid} {sid} {ym} = {got} (expect {exp})", fails)
    # the freeze re-mark pair (sweep 1c): Jan consumption +5.95, end-Jan working gas -137
    g_j, g_f = val("jan26", "NGWGPUS", "202601"), val("feb26", "NGWGPUS", "202601")
    fails = _t(g_j is not None and g_f is not None and abs((g_f - g_j) - (-137)) < 2.0,
               f"end-Jan working gas re-mark jan26 {g_j} -> feb26 {g_f} (~-137)", fails)
    # column-origin trap: detected per vintage, differs across the set
    fails = _t(sv["sep25"]["column_origin"] == "202101" and sv["jan26"]["column_origin"] == "202201",
               f"column origin detected (sep25 {sv['sep25']['column_origin']}, "
               f"jan26 {sv['jan26']['column_origin']}) - the named parse trap", fails)
    # blind wall: release day itself NOT knowable; release+1 is
    a = steo_vintage_asof("2026-01-13")
    fails = _t(a is not None and a["vintage_id"] == "dec25", f"2026-01-13 sees dec25 (got {a and a['vintage_id']})", fails)
    a = steo_vintage_asof("2026-01-14")
    fails = _t(a is not None and a["vintage_id"] == "jan26", f"2026-01-14 sees jan26 (got {a and a['vintage_id']})", fails)
    fails = _t(steo_vintage_asof("2025-09-09") is None, "2025-09-09 (sep25 release day) -> None", fails)
    fails = _t(steo_vintage_asof("2025-09-10")["vintage_id"] == "sep25", "2025-09-10 -> sep25", fails)
    # staleness anchors (sweep 1e): G11 open age 5, G12 open age 19
    fails = _t(steo_vintage_asof("2026-01-18")["age_days"] == 5, "G11 open (Jan 18) age 5d", fails)
    g12 = steo_vintage_asof("2026-02-01")
    fails = _t(g12["vintage_id"] == "jan26" and g12["age_days"] == 19, "G12 open (Feb 1) jan26 age 19d", fails)
    g12_tc = g12["fields"]["total_consumption_bcfd"]["prev"]
    fails = _t(g12_tc is not None and abs(g12_tc - 115.95) < 0.01,
               f"G12 open sees pre-freeze Jan consumption {g12_tc} (~115.95)", fails)
    # the re-mark becomes visible mid-G12 (Feb 11), as a REVISION the agent can read
    d11 = steo_vintage_asof("2026-02-11")
    rv = (d11 or {}).get("revisions_vs_prev_vintage") or {}
    tc = rv.get("total_consumption_bcfd") or {}
    fails = _t(d11 is not None and d11["vintage_id"] == "feb26" and tc.get("prev") is not None
               and abs(tc["prev"] - 5.95) < 0.02,
               f"2026-02-11 feb26 revision: Jan consumption delta {tc.get('prev')} (~+5.95)", fails)
    wg = rv.get("working_gas_inventory_bcf") or {}
    fails = _t(wg.get("prev") is not None and abs(wg["prev"] - (-137)) < 2.0,
               f"2026-02-11 feb26 revision: end-Jan working gas delta {wg.get('prev')} (~-137)", fails)
    # every walked trade day (Nov 3 - Feb 27) resolves to a vintage with release strictly prior
    d = datetime.date(2025, 11, 3)
    bad = 0
    while d <= datetime.date(2026, 2, 27):
        if d.weekday() < 5 or d.weekday() == 6:
            r = steo_vintage_asof(d.isoformat())
            if r is None or not (r["release_date"] < d.isoformat()):
                bad += 1
        d += datetime.timedelta(days=1)
    fails = _t(bad == 0, f"blind-wall walk Nov 3 - Feb 27: {bad} violations (expect 0)", fails)
    print(f"=== selftest {'PASS' if fails == 0 else f'FAIL ({fails})'} ===")
    return 1 if fails else 0


def main() -> int:
    ap = argparse.ArgumentParser(description="STEO monthly vintage balance feed (feed T)")
    ap.add_argument("--build", action="store_true")
    ap.add_argument("--force", action="store_true")
    ap.add_argument("--selftest", action="store_true")
    ap.add_argument("--show", metavar="DATE")
    ap.add_argument("--list-ids", action="store_true")
    a = ap.parse_args()
    if a.build:
        build(force=a.force)
        return 0
    if a.selftest:
        return _selftest()
    if a.show:
        print(json.dumps(steo_vintage_asof(a.show), indent=1))
        return 0
    if a.list_ids:
        st = load_store() or {"vintages": {}}
        for vid, v in st["vintages"].items():
            print(f"{vid}: {sorted(v['series'])}")
        return 0
    ap.print_help()
    return 0


if __name__ == "__main__":
    sys.exit(main())
